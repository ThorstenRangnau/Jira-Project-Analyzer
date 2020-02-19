import argparse
import csv
import re
import os

from github import Github
from jira import JIRA, JIRAError
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

from utils.spinner import Spinner
from jiraprojects.java_projects import APACHE_JAVA_PROJECTS
from jiraprojects.jira_project import JiraProject

NO_COMMENTS = "NO COMMENTS"

APACHE_JIRA_SERVER = 'https://issues.apache.org/jira/'

'''
python githubjiraparser.py -gr apache/activemq in order to parse apache active mq

python githubjiraparser.py -jp Cassandra -gr apache/cassandra

TODOs:

1. Receive all issues from jira for project:

2. Parse all issues and track comment metrics

3. Store all information in a csv file
'''


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-gr", dest="gitrepo", required=False,
        help="github repository to parse")
    parser.add_argument(
        "-jp", dest="jiraproject", required=False,
        help="Jira project to parse")
    return parser.parse_args()


def parse_github(repository_title, prefix):
    github = Github()
    repository = github.get_repo(repository_title)
    pull_requests = repository.get_pulls()
    for pr in pull_requests:
        if prefix in pr.title:
            # note this is for the moment in order to increase performance
            return True
        # print(pr.title)
        # print(pr.state)
        # commits = pr.get_commits()
        # TODO: atm we require log in to parse commits because of rate limit of 60
        # TODO: check how many prs/commits have prefix included
        # TODO: store results in order to map issue num and pr/commit
    return False


def get_absolute_path(rel_path):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    script_dir = re.sub("/src", "", script_dir)
    created_path = "%s%s" % (script_dir, rel_path)
    return created_path


def create_project_spreadsheet(project_name, issues, comments_for_issue):
    wb = Workbook()
    ws = wb.active
    ws.title = project_name
    # set the width of the particular columns in order to provide convenient reading
    # First 3 columns are for issue id, key and summary
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 65
    # Column 4 and 5 are for comment id and body
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 100
    # date
    ws['A1'] = "Downloaded on"
    ws['A2'] = datetime.today().strftime("%A, %d. %B %Y %I:%M%p")
    # row titles
    ws['D4'] = "Comment ID"
    ws['E4'] = "Comment Body"
    # TODO: maybe area is not really necessary
    area = 4
    for idx, issue in enumerate(issues):
        ws.append([None, None, None, None, None])
        ws.append(['Issue ID', 'Issue Key', 'Issue Summary'])
        ws.append([issue.id, issue.key, re.sub("(.{95})", "\\1\n", issue.fields.summary, 0, re.DOTALL)])
        area += 3
        comments = comments_for_issue[issue.key]
        if len(comments) > 0:
            for comment in comments:
                ws.append([None, None, None, comment.id, re.sub("(.{95})", "\\1\n", comment.body, 0, re.DOTALL)])
                area += 1
        else:
            ws.append([None, None, None, NO_COMMENTS, NO_COMMENTS])
            area += 1
    tab = Table(displayName=re.sub(" ", "", project_name), ref="A4:E%d" % (20 + area))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    spreadsheet_path = "/spreadsheets/comments-%s.xlsx" % re.sub(" ", "-", project_name)
    path_for_spreadsheet = get_absolute_path(spreadsheet_path)
    wb.save(path_for_spreadsheet)


def add_row_to_metrics_csv(name, jira_project_metric):
    metric_path = "/issue-metrics/apache-project-issue-metrics.csv"
    path_to_metrics = get_absolute_path(metric_path)
    with open(path_to_metrics, 'a', newline='') as file:
        writer = csv.writer(file)
        metrics = jira_project_metric.metrics()
        writer.writerow([name,
                         jira_project_metric.prefix,
                         metrics["total"],
                         metrics["average"],
                         metrics["commented_issues"],
                         metrics["average_commented"],
                         metrics["min"],
                         metrics["max"],
                         metrics["percentage_without"]])


def parse_apache_jira_projects():
    jira = JIRA(APACHE_JIRA_SERVER)
    projects = jira.projects()
    java_projects = list()
    for p in projects:
        if p.name in APACHE_JAVA_PROJECTS:
            java_projects.append(p)
    metrics = dict()
    for java_project in java_projects:
        # Load all issues for a given java project
        issues = jira.search_issues("project = \"%s\"" % java_project.name, maxResults=1)
        total_issues = total = issues.total
        all_issues = list()
        start = 0
        print("Start loading %d issues for project %s" % (total_issues, java_project.name))
        while total_issues > 0:
            print("**** LOADING: Issue %d - %d" % (start, start + 999 if start + 999 < total else total))
            with Spinner():
                try:
                    issues = jira.search_issues("project = \"%s\"" % java_project.name, startAt=start, maxResults=1000)
                except JIRAError:
                    return
            all_issues.extend(issues)
            start += 1000
            total_issues -= 1000
        print("Project %s  has %d total issues" % (java_project.name, issues.total))
        comments_issue = dict()
        comments_by_issue = dict()
        # Load all comments for each issue
        print("Load comments for issues of project %s!" % java_project)
        with Spinner():
            for issue in all_issues:
                try:
                    comments = jira.comments(issue)
                    comments_by_issue[issue.key] = comments
                    comments_issue[issue.key] = len(comments)
                except JIRAError:
                    return
        create_project_spreadsheet(java_project.name, issues, comments_by_issue)
        issue_name = all_issues[0].key
        issue_prefix = re.sub("\d+", "", issue_name)
        issue_prefix = re.sub("-", "", issue_prefix)
        add_row_to_metrics_csv(java_project.name, JiraProject(java_project.name, issue_prefix, total, comments_issue))
        metrics[java_project.name] = JiraProject(java_project.name, issue_prefix, total, comments_issue)
    # i = 0
    # for issue in issues:
    #     i += 1
    #     issue_name = issue.key
    # issue_prefix = re.sub("\d+", "", issue_name)
    # issue_prefix = re.sub("-", "", issue_prefix)
    # print(i)
    # print("Found %d java projects on Jira out of %d apache java projects" % (
    #     len(java_projects), len(APACHE_JAVA_PROJECTS)))
    # for k, v in metrics.items():
    #     print(k)
    #     for a, b in v.comments.items():
    #         print("%s with %d comments" % (a, b))
    return metrics


if __name__ == "__main__":
    args = parse_args()
    jira_issue_metrics = parse_apache_jira_projects()
